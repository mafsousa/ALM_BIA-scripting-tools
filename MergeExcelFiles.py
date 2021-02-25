## Copy Data Columns From Excel to Excel file
## Version:v0.1
##
## Short description: Input two excel (.csv) files (source and target) and the columns number to copy from source to target file; 
##                  leave columns number empty if you want to copy all the columns;
##                  save source changed file as merged.xls

## Requires python 3.6, openpyxl, pysimplegui ('python -m pip install openpyxl', 'python -m pip install pysimplegui')
## 
## This script should NOT be redistributed without author's permission. 
## Explicit acknowledgement to the ALM facility should be done in case of published articles (approved in C.E. 7/17/2017):     
## 
## "The authors acknowledge the support of i3S Scientific Platform Advanced Light Microscopy, 
## member of the national infrastructure PPBI-Portuguese Platform of BioImaging (supported by POCI-01-0145-FEDER-022122)."
## 
## Date: February/2021
## Author: Mafalda Sousa, mafsousa@ibmc.up.pt 
## Advanced Ligth Microscopy, I3S 
## PPBI-Portuguese Platform of BioImaging
##



# import modules
import os,csv, sys
import openpyxl as xl;
from openpyxl import Workbook
import PySimpleGUI as sg

# create GUI
sg.change_look_and_feel('Dark2') 

file_list_column = [
    [
        sg.Text("Input file 1"),
        sg.In(size=(25, 1), enable_events=True, key="-FOLDER1-"),
        sg.FileBrowse(),
    ],
    [
        sg.Text("Input file 2"),
        sg.In(size=(25, 1), enable_events=True, key="-FOLDER2-"),
        sg.FileBrowse(),
    ],
    [
        sg.Text(text='Define excel columns number'), sg.InputText('', size=(10, 1), key='-cols-'),
    ],
    [
        sg.Button('Ok'), sg.Button('Cancel'),
    ]
]
layout = [
    [
        sg.Column(file_list_column)       
    ]
]

window = sg.Window("Merge excel files", layout)

# Create an event loop
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel': # if user closes window or clicks cancel
        window.close()
        sys.exit(0)
    elif event == 'Ok':
        break

window.close()
if(values['-cols-'] != ""):
    columns_str = values['-cols-'].split(",")
    try:
        columns = [int(i) for i in columns_str]
    except:
        print("Invalid columns number")
        sys.exit(0)
else:
    columns = []

   
def process_sequence(file1, file2, columns):
	
            dirname = os.path.dirname(file1)
            #print(dirname)
            
            # opening the source excel file 
            
            wb1 = Workbook()
            ws1 = wb1.active
            with open(file1, 'r') as f:
                for row in csv.reader(f):
                    ws1.append(row)

            # opening the destination excel file 
            wb2 = Workbook()
            ws2 = wb2.active
            with open(file2, 'r') as f:
                for row in csv.reader(f):
                    ws2.append(row)
                    
            # calculate total number of rows and 
            # columns in source excel file 
            mr = ws1.max_row 
            mc = ws1.max_column 

            #columns = [3,7,8]
            if (columns == []): #if columns input is empty, select all
                columns =  list(range(1,mc))
            if (all (x < mc for x in columns)==False):
                print("Invalid columns number")
                sys.exit(0)
    
                
            # copying the cell values from source 
            # excel file to destination excel file 
            for i in range (1, mr + 1):
                    col = 0
                    for j in columns:
                            col = col + 1
                            # reading cell value from source excel file 
                            c = ws1.cell(row = i, column = j) 

                            # writing the read value to destination excel file
                            if i == 1: # change headers names
                                ws2.cell(row = i, column = mc+col).value = c.value + "_file1"
                            else:
                                ws2.cell(row = i, column = mc+col).value = c.value  
       
            # saving the destination excel file 
            wb2.save(str(dirname + os.path.sep + "merged.xls")) 


process_sequence(values['-FOLDER1-'],values['-FOLDER2-'],columns)
print("===Merged done===")

