##Title: InCell Merge Excel Files
##Version:v0.1
##
## Short description: Input the InCell results folder with at least 2 .csv files (results only from green, red channels); 
##                  copy three columns (relative to intesity values) from green file and add it to red file;
##                  save changed red channel file in .xls
## Prerequisites, python > 3.7 openpyxl
## install openpyxl (pip3 install openpyxl)
## 
## This macro should NOT be redistributed without author's permission. 
## Explicit acknowledgement to the ALM facility should be done in case of published articles (approved in C.E. 7/17/2017):     
## 
## "The authors acknowledge the support of i3S Scientific Platform Advanced Light Microscopy, 
## member of the national infrastructure PPBI-Portuguese Platform of BioImaging (supported by POCI-01-0145-FEDER-022122)."
## 
## Date: April/2020
## Author: Mafalda Sousa, mafsousa@ibmc.up.pt 
## Advanced Ligth Microscopy, I3S 
## PPBI-Portuguese Platform of BioImaging
##

# importing openpyxl module 
import openpyxl as xl;
from openpyxl import Workbook
import csv
import os
from glob import glob
from tkinter import Tk
from tkinter.filedialog import askdirectory

root = tk.Tk()
path = askdirectory(title='Select Folder') # shows dialog box and return the path
print(path)
root.withdraw()
#directory = "C:\\Users\\ALM\\Downloads\\Ficheiros Excel\\Experiment3\\Green signal"
extension = ".csv"
process_folder(path, extension) 

def process_sequence(folder, files):
	
            dirname = folder
            print(dirname)
            for f in range(0,len(files)):
                filename_in = dirname + "\\" + os.path.basename(files[f])
                dirname_red = dirname.replace("Green","Red")
                filename_red = os.path.basename(files[f]).replace("_Green","")
                filename_out = dirname_red + "\\" + filename_red
                print(filename_in)
                print(filename_out )
		
            # opening the source excel file 
            
            wb1 = Workbook()
            ws1 = wb1.active
            with open(filename_in, 'r') as f:
                for row in csv.reader(f):
                    ws1.append(row)

            # opening the destination excel file 
            wb2 = Workbook()
            ws2 = wb2.active
            with open(filename_out, 'r') as f:
                for row in csv.reader(f):
                    ws2.append(row)
                    
            # calculate total number of rows and 
            # columns in source excel file 
            mr = ws1.max_row 
            mc = ws1.max_column 

            columns = [3,7,8]
            # copying the cell values from source 
            # excel file to destination excel file 
            for i in range (1, mr + 1):
                    col = 0
                    for j in columns:
                            col = col + 1
                            # reading cell value from source excel file 
                            c = ws1.cell(row = i, column = j) 

                            # writing the read value to destination excel file
                            if i == 1: # change headers names
                                ws2.cell(row = i, column = mc+col).value = c.value + "_Green"
                            else:
                                ws2.cell(row = i, column = mc+col).value = c.value 
       
            # saving the destination excel file 
            wb2.save(str(filename_out)) 


def process_folder(folder, extension):	
	for subdir in glob(os.path.join(folder, '*/')):
		process_folder(subdir, extension)
		
		files = glob(os.path.join(folder, '*' + extension))
		if len(files) > 0:
				process_sequence(folder, files)

		   

       


